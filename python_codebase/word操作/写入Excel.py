import openpyxl
import re

# 读取excel文件，获取workbook对象
wb = openpyxl.Workbook()  # 创建Workbook()对象
ws = wb.active  # 获取默认工作薄
wb.save("./HCIA-B.xlsx")  # 保存


data = open('./HCIA-B.txt', encoding='utf-8').readlines()
wb = openpyxl.load_workbook("./HCIA-B.xlsx")
sheet = wb['Sheet']  # 通过名称获取工作薄

list = []
aa = []

for i in data:
    if i != '\n':
        aa.append(i.replace('\n', ''))
    else:
        list.append(aa)
        aa = []

# print(list)
#
for i in list:
    # print(i[0].isalnum())
    # aa = re.split(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])", i[0])
    aa = re.split("[\(\（\）\)\[\]\.]", i[0])
    i[0] = aa[0]

    for j in aa:
        if j != '' and j != ' ':
            i.insert(0, j)

    aaa = ['概念理解',
           i[0],
           i[1],
           i[2]]
    for c in i:
        if '.' in c:
            aaa.append(c)
    sheet.append(aaa)  # 插入一行数据

wb.save("./HCIA-B.xlsx")  # 保存,传入原文件则在原文件上追加数据，也可以保存为新文
