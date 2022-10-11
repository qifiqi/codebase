# -*- coding: utf-8 -*-
# All Rights Reserved 
# @Time    : 2022/7/2211:37
# @Author  : 小符玩代码
# @Email   : 2737454073@qq.com
# @File    : kaoshibao.py
# @Software: PyCharm
import json
import pprint
import time
import execjs
import re
import queue
import requests
from faker import Faker
from openpyxl import Workbook

fak = Faker(locale='zh_CN')
row_queue = queue.Queue()
file = ''

# 题目类型
topic_type_dict = {
    '1': '单选题',
    '2': '多选题',
    '3': '判断题',
    '4': '填空题',
    '15': '排序题',
    '14': '不定项选择题',
    '5': '简答题',
    '9': '案例分析',
    '11': '论述题',
    '12': '计算题',
}

head = {
    'user-agent': fak.chrome(),
    'Accept': "application/json, text/plain, */*",
    'CLIENT-IDENTIFIER': "",
    'Content-Type': "application/json;charset=utf-8",
    'device': "zaixiankaoshi",
    'Sign': "460d9702d88e887dbd0127374cd55e4f",
    'TimeStamp': '1658467885823',
    'platform': "web",
    'cookie': ""
}
"""爬取设置"""
"""题库编号"""
paperid = "861559" #国家公务员-行政能力测试
# paperid = "6318639"  # HCIP大数据.json
# paperid = "4081241"
head[
    'cookie'] = 'UM_distinctid=1824f37b4ea24a-01c372a654161c-76492e29-144000-1824f37b4ebbca; __gads=ID=a584ec7ee881aac2-22a692bc61d50055:T=1659186428:RT=1659186428:S=ALNI_MZm1eNZ5lrPwtEQ2kuait6xCOwMQw; __gpi=UID=00000821f06e05c9:T=1659186428:RT=1659186428:S=ALNI_MZRyqMTDf0KcA52ncSM0ce2dAHFXg; CNZZDATA1278923901=951862156-1659184477-https%253A%252F%252Fwww.zaixiankaoshi.com%252F%7C1659188081; Hm_lvt_fafd23c81b939a6c401dc831cc0eeac6=1659179142,1659186428,1659189221,1659190134; uu=376521f8-8c6d-48cc-8fab-1d794a4b427b; token=icgBFsyOIRW%2F6KLNPJSwP4nKR%2FHjDD1pEQvFVlkJsx4%3D; SOTOKEN=icgBFsyOIRW%2F6KLNPJSwP4nKR%2FHjDD1pEQvFVlkJsx4%3D; Hm_lpvt_fafd23c81b939a6c401dc831cc0eeac6=1659190137'


def sign(api_url: str) -> str:
    """
    sign加密参数
    :param api_url: 网站来好api
    :return: 返回加密数据
    """
    o = "12b6bb84e093532fb72b4d65fec3f00b"  # 不变
    # t = "c31815dc-d4af-4189-838a-ee8c1690e27c"  # cookie里面的一个uu
    t = head['CLIENT-IDENTIFIER']  # cookie里面的一个uu
    c = api_url.split('api')[-1]  # api地址
    n = int(time.time() * 1000)  # 时间戳
    cc = (o + t + c + str(n) + o)

    node = execjs.get()
    fanke = node.compile(open('./kab_md5.js', encoding='utf-8').read())

    aa = 'ps()("{t}","{e}")'.format(t=cc, e='hex')

    return fanke.eval(aa), str(n)


def dict_cookie(cookie_str: str) -> dict:
    """
    用于获取参数CLIENT-IDENTIFIER
    :param cookie_str: 吧cookie传入
    :return: 返回解析的cookie
    """
    # cookie_str = 'UM_distinctid=180e1ecf7c0a0c-0a0f4c16d3b597-4c647e56-144000-180e1ecf7c141c; __gads=ID=47131fea20bb1962-227ddad8d0d3008f:T=1654576886:RT=1654576886:S=ALNI_MarOK2d8NVYRwMYuJfBhCTxbPjWIA; __gpi=UID=0000066e79c927c4:T=1654576886:RT=1658459397:S=ALNI_MZ44apm0D28iv82jjrjPtCvK7rV2w; uu=c31815dc-d4af-4189-838a-ee8c1690e27c; Hm_lvt_fafd23c81b939a6c401dc831cc0eeac6=1658459397,1658460132; CNZZDATA1278923901=2131432509-1654575948-https%253A%252F%252Fwww.zaixiankaoshi.com%252F%7C1658466970; Hm_lpvt_fafd23c81b939a6c401dc831cc0eeac6=1658466971'
    cookie = {}
    for i in re.split(';|:', cookie_str):
        i_list = i.split('=', 1)
        if len(i_list) > 2:
            continue
        key, values = i_list
        key = key.strip()
        values = values.strip()

        cookie[key] = values
    return cookie


def fetch() -> json:
    """
    获取所有题目id
    :return:返回所有题目的json
    """
    url = 'https://www.zaixiankaoshi.com/api/questions/fetch'
    head['Sign'], head['TimeStamp'] = sign(url)

    data = {
        'paperid': paperid,
        'qtype': "",
        'chapter': "0",
        'size': "30000"
    }

    response = requests.post(url, headers=head, data=data)
    if response.status_code == 200:
        res_json = response.json()
        return res_json


def ids(row: str) -> json:
    """
    返回题目
    :rtype: json 返回10个题题目的详情
    """
    url = 'https://www.zaixiankaoshi.com/api/questions/ids'
    head['Sign'], head['TimeStamp'] = sign(url)

    data = {
        "ids": row,
        "paperid": paperid
    }
    print(data)
    response = requests.post(url, headers=head, data=data)
    if response.status_code == 200:
        res_json = response.json()
        Parse_ids(res_json)


def Parse_ids(res_json: json) -> None:
    """
    解析题目
    :param res_json: 传入题目json
    """
    # res_json = json.loads(open('./ids.json', 'r', encoding='utf-8').read())
    for data in res_json['data']:
        data_dict = {
            'id': data['id'],
            'paperid': data['paperid'],
            'question': data['question'],
            'answer': data['answer'],
            'topic_type': topic_type_dict[data['qtype']],
            'datatime': data['created_at'],
            'options': []
        }
        options = data['options']
        if len(options) > 0:
            for i in json.loads(options):
                data_dict['options'].append({i['Key']: i['Value']})
        file.write(json.dumps(data_dict, ensure_ascii=False) + ',\n')


def start_crawl():
    """爬取程序开始"""
    # 初始化cookis中的uu,赋值到请求头中的CLIENT-IDENTIFIER里
    head['CLIENT-IDENTIFIER'] = dict_cookie(cookie_str=head['cookie'])['uu']
    # 返回ids_row的json
    fe_row = fetch()
    file_name = re.sub('[^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a]', '', fe_row["data"]["paper"]['name'])

    global file
    file = open(f'./{file_name}.json', 'a+', encoding='utf-8')
    file.write('[\n')
    if fe_row:
        # 处理ids_fetch
        row_list = fe_row['data']['rows']
        len_row = int(len(row_list) / 10)
        q = 0
        for i in range(1, len_row + 1):
            ids(json.dumps(row_list[q:i * 10]))
            q = i * 10
        else:
            ids(json.dumps(row_list[q:]))
    file.write('\n]')
    file.flush()
    file.close()
    xlsx_create(file_name)


def xlsx_create(title):
    """
    将json文件转换为考试宝的xlsx文件
    :param title:
    """
    json_file = json.loads(open(f'./{title}.json', 'r', encoding='utf-8').read())
    wb = Workbook()  # 新建Excel工作簿
    ws = wb.active  # 使用活动工作表
    ws.cell(1, column=1, value='题干（必填)')
    ws.cell(1, column=2, value='题型 （必填）')
    ws.cell(1, column=3, value='选项 A')
    ws.cell(1, column=4, value='选项 B')
    ws.cell(1, column=5, value='选项 C')
    ws.cell(1, column=6, value='选项 D')
    ws.cell(1, column=7, value='选项E(勿删)')
    ws.cell(1, column=8, value="选项F(勿删)")
    ws.cell(1, column=9, value="选项G(勿删)")
    ws.cell(1, column=10, value="选项H(勿删)")
    ws.cell(1, column=11, value="正确答案（必填）")
    ws.cell(1, column=12, value='解析')
    ws.cell(1, column=13, value='章节')
    ws.cell(1, column=14, value='难度')

    for row, data in enumerate(json_file):
        row = row + 2
        ws.cell(row, column=1, value=data['question'])
        ws.cell(row, column=2, value=data['topic_type'])
        for i in range(3, 11):
            if i - 3 < len(data['options']):
                ws.cell(row, column=i, value=list(data['options'][i - 3].values())[0])
            else:
                ws.cell(row, column=i, value='')
        ws.cell(row, column=11, value=data['answer'])
        ws.cell(row, column=12, value='')
        ws.cell(row, column=13, value='')
        ws.cell(row, column=14, value='')

    wb.save(f'./{title}.xlsx')
    pass


def main():
    """
    主程序
    """

    start_crawl()


if __name__ == '__main__':
    main()
