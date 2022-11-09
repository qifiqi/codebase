# -*- coding: utf-8 -*-
# All Rights Reserved 
# @Time    : 2022/9/2713:30
# @Author  : Small Fu
# @Email   : 2737454073@qq.com
# @File    : get_page_list.py
__author__ = 'Small Fu'

import json
import pprint
import time
from typing import Tuple, Any

import execjs
import re
import queue
import requests
import pandas as pd
from faker import Faker
from openpyxl import Workbook
from urllib.parse import quote

fak = Faker(locale='zh_CN')
row_queue = queue.Queue()

head = {
    'user-agent': fak.chrome(),
    'Accept': "application/json, text/plain, */*",
    'CLIENT-IDENTIFIER': "",
    'Content-Type': "application/json;charset=utf-8",
    'device': "zaixiankaoshi",
    'Sign': "d006208f1938e7d543d1f88f74c38b6e",
    'TimeStamp': '1658467885823',
    'platform': "web",
    'cookie': input("cookie:"),
    "authorization": "%22V2rhmypP5ORpQyiTBkhyDInKR/HjDD1pEQvFVlkJsx4=%22"
}


def dict_cookie(cookie_str: str) -> dict:
    """
    用于获取参数CLIENT-IDENTIFIER
    :param cookie_str: 吧cookie传入
    :return: 返回解析的cookie
    """
    cookie = {}
    for i in re.split(';|:', cookie_str):
        i_list = i.split('=', 1)
        if len(i_list) > 2:
            continue
        key, values = i_list
        key = key.strip()
        values = values.strip()

        cookie[key] = values
    return cookie


# head['cookie'] = "UM_distinctid=1824f37b4ea24a-01c372a654161c-76492e29-144000-1824f37b4ebbca; __gads=ID=a584ec7ee881aac2-22a692bc61d50055:T=1659186428:RT=1659186428:S=ALNI_MZm1eNZ5lrPwtEQ2kuait6xCOwMQw; __gpi=UID=00000821f06e05c9:T=1659186428:RT=1661921629:S=ALNI_MZRyqMTDf0KcA52ncSM0ce2dAHFXg; Hm_lvt_fafd23c81b939a6c401dc831cc0eeac6=1662178574,1662271247,1664255984,1664256790; CNZZDATA1278923901=2017846672-1664285189-https%253A%252F%252Fwww.zaixiankaoshi.com%252F%7C1664285189; uu=e321d8ae-d0ff-4c32-9d6e-4d3fa22f2105; token=V2rhmypP5ORpQyiTBkhyDInKR%2FHjDD1pEQvFVlkJsx4%3D; SOTOKEN=V2rhmypP5ORpQyiTBkhyDInKR%2FHjDD1pEQvFVlkJsx4%3D; Hm_lpvt_fafd23c81b939a6c401dc831cc0eeac6=1664287059"

dict_cookies = dict_cookie(cookie_str=head['cookie'])
head['CLIENT-IDENTIFIER'] = dict_cookies['uu']
head['authorization'] = quote("\"")+dict_cookies.get('token')+quote("\"")

def sign(api_url: str) -> tuple[Any, str]:
    """
    sign加密参数
    :param api_url: 网站来好api
    :return: 返回加密数据
    """
    o = "12b6bb84e093532fb72b4d65fec3f00b"  # 不变
    # t = "c31815dc-d4af-4189-838a-ee8c1690e27c"  # cookie里面的一个uu
    t = head['CLIENT-IDENTIFIER']  # cookie里面的一个uu
    c = api_url.split('api')[-1]  # api地址
    n = int(time.time() * 1000)  # 时间戳
    cc = (o + t + c + str(n) + o)

    node = execjs.get()
    fanke = node.compile(open('./kab_md5.js', encoding='utf-8').read())

    aa = 'ps()("{t}","{e}")'.format(t=cc, e='hex')

    return fanke.eval(aa), str(n)


def exam_page(data):
    url = "https://www.zaixiankaoshi.com/api/user/exam/page"

    head['Sign'], head['TimeStamp'] = sign(url)
    head['referer'] = url
    response = requests.post(url, headers=head, data=data)
    try:
        res_json = response.json()
        res_json_row = res_json.get("data").get('rows')
        for res in res_json_row:
            examResult_getRankingList(exa_id=res.get('id'))
        # examResult_getRankingList(exa_id=res_json_row[1].get('id'))
    except Exception as ex:
        print(ex)



def examResult_getRankingList(exa_id: object = '3756895') -> object:
    """
    用来请求
    :param exa_id:
    """

    def sub_symbol(strs):
        com = re.compile('[^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a]')
        return com.sub('', strs)

    url = "https://www.zaixiankaoshi.com/api/user/examResult/page"
    data = {"id": f"{exa_id}", "page": 1, "keyword": "", "size": 20, "column": "", "direction": ""}
    head['Sign'], head['TimeStamp'] = sign(url)
    head['referer'] = url
    response_page = requests.post(url, headers=head, data=data)

    url = "https://www.zaixiankaoshi.com/api/user/examResult/getRankingList"
    head['Sign'], head['TimeStamp'] = sign(url)
    data = {"id": f"{exa_id}", "size": 500}
    head['referer'] = url
    response = requests.post(url, headers=head, data=data)
    try:
        res_json = response.json()
        title = sub_symbol(response_page.json().get('data').get("exam").get("title"))
        print(title)
        res_json_row = res_json.get("data").get('rows')

    except Exception as ex:
        print(ex)
        return ""
    if res_json_row is not None:
        res_particulars = []
        for res in res_json_row:
            login_options = json.loads(res.get("login_options"))
            res_particulars.append({
                "stu_name": login_options[0].get("value"),
                "phone": login_options[1].get("value"),
                "passed": res.get("passed"),
                "score": res.get("score"),
            })
        pd_exa(res_particulars, title)


def pd_exa(res_particulars, title='11'):
    def size_num(x, num):
        if int(x) >= num:
            return 1
        else:
            return 0

    json_data = pd.DataFrame(res_particulars).sort_values(by="score", ascending=False)
    json_data["score"] = json_data["score"].map(lambda x: int(x.split(".")[0]))
    json_data['passed'] = json_data['score'].map(lambda x: size_num(x=x, num=60))
    # 平均数
    average_score = json_data["score"].mean()
    # 通过率
    passing_rate = (json_data[json_data['passed'] == 1]["passed"].count() / json_data['passed'].count()) * 100
    exl_save(json_data, average_score, passing_rate, title)


def exl_save(json_data, average_score, passing_rate, title):
    wb = Workbook()  # 新建Excel工作簿
    ws = wb.active  # 使用活动工作表
    for i, index in enumerate(json_data.columns):
        ws.cell(1, column=i + 1, value=index)

    index = 1
    for val in json_data.values:
        index += 1
        for i, v in enumerate(val):
            ws.cell(index, column=i + 1, value=v)
    else:
        index += 1
    ws.cell(index, column=1, value="平均分")
    ws.cell(index, column=2, value="合格率")
    index += 1
    ws.cell(index, column=1, value=average_score)
    ws.cell(index, column=2, value=passing_rate)

    wb.save(f'./{title}.xlsx')


if __name__ == '__main__':
    data = {"keyword": "", "page": 1, "size": 10, "show_daily": "1"}
    exam_page(data)  # 入口直接获取最新
    data = {"keyword": "", "page": 2, "size": 10, "show_daily": "1"}
    exam_page(data)  # 入口直接获取最新

    # examResult_getRankingList('3756927')
    # examResult_getRankingList('3756895')
    # pd_exa()
