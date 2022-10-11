# -*- coding: utf-8 -*-
# All Rights Reserved 
# @Time    : 2022/7/3021:55
# @Author  : Small Fu
# @Email   : 2737454073@qq.com
# @File    : execl操作.py
__author__ = 'Small Fu'

import json

from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment  # 字体，颜色，对齐
from openpyxl.styles import PatternFill  # 单元格填充

def xlsx_create(title):
    json_file = json.loads(open(f'./{title}.json', 'r', encoding='utf-8').read())
    wb = Workbook()  # 新建Excel工作簿
    ws = wb.active  # 使用活动工作表
    ws.cell(1, column=1, value='题干（必填)')
    ws.cell(1, column=2, value='题型 （必填）')
    ws.cell(1, column=3, value='选项 A')
    ws.cell(1, column=4, value='选项 B')
    ws.cell(1, column=5, value='选项 C')
    ws.cell(1, column=6, value='选项 D')
    ws.cell(1, column=7, value='选项E(勿删)')
    ws.cell(1, column=8, value="选项F(勿删)")
    ws.cell(1, column=9, value="选项G(勿删)")
    ws.cell(1, column=10, value="选项H(勿删)")
    ws.cell(1, column=11, value="正确答案（必填）")
    ws.cell(1, column=12, value='解析')
    ws.cell(1, column=13, value='章节')
    ws.cell(1, column=14, value='难度')

    for row, data in enumerate(json_file):
        row = row + 2
        ws.cell(row, column=1, value=data['question'])
        ws.cell(row, column=2, value=data['topic_type'])
        for i in range(3, 11):
            if i - 3 < len(data['options']):
                ws.cell(row, column=i, value=list(data['options'][i - 3].values())[0])
            else:
                ws.cell(row, column=i, value='')
        ws.cell(row, column=11, value=data['answer'])
        ws.cell(row, column=12, value='')
        ws.cell(row, column=13, value='')
        ws.cell(row, column=14, value='')

    wb.save(f'./{title}.xlsx')
    pass



if __name__ == '__main__':
    xlsx_create('HCIP大数据')
